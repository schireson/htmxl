"""Module dedicated to adding, finding, and applying styles."""
import functools
import logging
from copy import copy

from openpyxl.styles import (
    Alignment,
    Border,
    Fill,
    Font,
    GradientFill,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)

logger = logging.getLogger(__name__)


default_styles = [
    {"name": "xl-centered", "alignment": {"horizontal": "center"}},
    {"name": "xl-underlined", "font": {"underline": "single"}},
]


class Styler:
    def __init__(self, wb, styles=None):
        self.wb = wb
        self.named_styles = {}
        self.named_style_specs = {}

        self.register_styles(default_styles)

        if styles is not None:
            self.register_styles(styles)

    def get_style(self, element):
        return element.classes

    def get_inline_style(self, element):
        inline_style = {}
        raw_inline_style = element.get("style", None)
        data_attributes = {
            attr: value for attr, value in element.attrs.items() if attr.startswith("data-")
        }
        inline_style.update(data_attributes)
        if raw_inline_style is None:
            return inline_style

        for inline_style_el in raw_inline_style.split(";"):
            if inline_style_el:
                name, value = inline_style_el.split(":")

                if name in {"width", "min-width", "max-width"}:
                    value = int(value[0 : value.index("ch")])

                elif name in {"height"}:
                    value = int(value[0 : value.index("px")])

                if isinstance(value, str):
                    value = value.strip()

                inline_style[name] = value

        return inline_style

    def register_style(self, style):
        name = style["name"]
        logger.debug("Registering named style {} in workbook {}".format(name, self.wb))

        named_style = make_style(style)

        self.wb.add_named_style(named_style)
        self.named_styles[name] = named_style
        self.named_style_specs[name] = style

    def register_styles(self, styles):
        for style in styles:
            self.register_style(style)

    def calculate_style(self, styles):
        if not styles:
            return None

        if len(styles) == 1:
            return styles[0]

        name = "_".join(style for style in styles)

        # Normally duplicate `register_style` calls for a given name intentionally cause an error,
        # but given that this is dynamic we dont want to regen or error on subsequent usages.
        if name not in self.named_style_specs:
            style_specs = [self.named_style_specs[style] for style in styles]
            result = functools.reduce(lambda a, b: dict_merge(a, b), style_specs, {})
            result["name"] = name
            self.register_style(result)
        return name


def make_style(style_spec):
    name = style_spec["name"]
    style = NamedStyle(name)

    if "font" in style_spec:
        style.font = Font(**style_spec["font"])

    if "border" in style_spec:
        border_spec = copy(style_spec["border"])

        for side in {"top", "bottom", "left", "right"}:
            if side in style_spec["border"]:
                border_spec[side] = Side(**style_spec["border"][side])

        style.border = Border(**border_spec)

    if "alignment" in style_spec:
        style.alignment = Alignment(**style_spec["alignment"])

    if "pattern_fill" in style_spec:
        style.fill = PatternFill(**style_spec["pattern_fill"])
    elif "gradient_fill" in style_spec:
        style.fill = GradientFill(**style_spec["gradient_fill"])
    elif "fill" in style_spec:
        style.fill = Fill(**style_spec["fill"])

    if "number_format" in style_spec:
        style.number_format = style_spec["number_format"]

    if "protection" in style_spec:
        style.protection = Protection(**style_spec["protection"])

    return style


def style_range(ws, reference_style, ref):
    top = Border(top=reference_style.border.top)
    left = Border(left=reference_style.border.left)
    right = Border(right=reference_style.border.right)
    bottom = Border(bottom=reference_style.border.bottom)

    first_cell = ws[ref.split(":")[0]]
    if reference_style.alignment:
        ws.merge_cells(ref)
        first_cell.alignment = reference_style.alignment

    rows = ws[ref]
    if reference_style.font:
        first_cell.font = reference_style.font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        left_cell = row[0]
        right_cell = row[-1]
        left_cell.border = left_cell.border + left
        right_cell.border = right_cell.border + right
        if reference_style.fill:
            for column in row:
                column.fill = reference_style.fill


def dict_merge(d, other, add_keys=True):
    d = d.copy()
    for k, v in other.items():
        if k in d and isinstance(d[k], dict) and isinstance(other[k], dict):
            d[k] = dict_merge(d[k], other[k])
        else:
            d[k] = other[k]

    return d
