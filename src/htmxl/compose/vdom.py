from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl.worksheet.datavalidation import DataValidation


@dataclass
class VDom:
    landscape: Dict[Tuple[int, int], Cell] = field(default_factory=dict)

    cursor_x: int = 0
    cursor_y: int = 0

    width: int = 1
    height: int = 1

    ancestors: List[Cell] = field(default_factory=list)

    @property
    def cursor(self):
        return (self.cursor_x, self.cursor_y)

    def cell(self):
        cell = self.landscape.get(self.cursor)
        if cell is None:
            ancestor = self.ancestors[-1] if self.ancestors else None
            cell = Cell(ancestor=ancestor)
            self.landscape[self.cursor] = cell
        return cell

    def move_left(self):
        if self.cursor_x > 0:
            self.cursor_x -= 1
        self.cell()

    def move_right(self):
        self.cursor_x += 1

        if self.cursor_x == self.width:
            self.width += 1
        self.cell()

    def move_up(self):
        if self.cursor_y > 0:
            self.cursor_y -= 1
        self.cell()

    def move_down(self):
        self.cursor_y += 1

        if self.cursor_y == self.height:
            self.height += 1
        self.cell()

    def move_to(self, x, y):
        self.cursor_x = x
        self.cursor_y = y
        self.cell()

    def write(self, **options):
        cell = self.cell()
        cell.write(**options)

    def enter_context(self):
        cell = self.cell()
        self.ancestors.append(cell)

    def exit_context(self):
        if self.ancestors:
            self.ancestors.pop(-1)

    def rows(self):
        for row_index in range(self.height):
            row = [self.landscape[(column_index, row_index)] for column_index in range(self.width)]
            yield row


@dataclass
class Cell:
    ancestor: Optional[Cell] = None

    self_styles: List[Style] = field(default_factory=list)
    validations: List[DataValidation] = field(default_factory=list)
    text: str = ""

    @property
    def styles(self):
        if self.ancestor:
            yield from self.ancestor.styles
        yield from self.self_styles

    def write(
        self,
        ancestor=None,
        style=None,
        text=None,
        validation=None,
    ):
        if ancestor is not None:
            self.ancestor = ancestor

        if style is not None:
            self.self_styles.append(style)

        if text is not None:
            self.text = text

        if validation is not None:
            self.validations.append(validation)


@dataclass
class Style:
    background_color: Optional[str] = None
    font_color: Optional[str] = None
