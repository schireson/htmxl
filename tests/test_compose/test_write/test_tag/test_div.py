FIXTURE_DIR = "tests/fixtures/templates/tags/div"


class WriteBRTests:

    template_file = ""
    expected_result_file = ""

    @property
    def expected_result(self):
        import openpyxl
        import os

        return openpyxl.load_workbook(os.path.join(FIXTURE_DIR, self.expected_result_file))

    @property
    def result(self):
        import os
        import openpyxl
        from schireson_excel.compose import Workbook
        from schireson_excel.compose.compat import BytesIO

        template_file = os.path.join(FIXTURE_DIR, self.template_file)
        wb = Workbook()
        wb.add_sheet_from_template_file(template_file)
        fileobj = BytesIO()
        wb.compose(fileobj)
        fileobj.seek(0)
        return openpyxl.load_workbook(fileobj)

    def test_sheets_num(self):
        assert len(self.expected_result.worksheets) == len(self.result.worksheets)

    def test_equality(self):
        expected_worksheet = self.expected_result.worksheets[0]
        result_worksheet = self.result.worksheets[0]

        for result_row, expected_row in zip(result_worksheet.rows, expected_worksheet.rows):
            for result_cell, expected_cell in zip(result_row, expected_row):
                assert result_cell.value == expected_cell.value
                assert result_cell.style == expected_cell.style


class TestAdjacentDivs(WriteBRTests):
    template_file = "adjacent_divs.html.jinja2"
    expected_result_file = "adjacent_divs.xlsx"


class TestContainedDivs(WriteBRTests):
    template_file = "contained_divs.html.jinja2"
    expected_result_file = "contained_divs.xlsx"


class TestNoContent(WriteBRTests):
    template_file = "div_no_content.html.jinja2"
    expected_result_file = "div_no_content.xlsx"
