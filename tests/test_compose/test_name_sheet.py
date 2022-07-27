import io

import openpyxl

from htmxl.compose import Workbook

data = dict(
    title="Hello World",
    name="Bob",
    column_names=["A", "B"],
    rows=[{"a": str(i), "b": int(i)} for i in range(10000)],
)


def test_name_sheet():
    with open("tests/test_performance.jinja2") as f:
        template = f.read()

    SHEET_NAME = "Test Foo 123"
    workbook = Workbook(parser="lxml")
    workbook.add_sheet_from_template(template=template, data=data, sheet_name=SHEET_NAME)

    buffer = io.BytesIO()
    workbook.compose(buffer)

    wb = openpyxl.load_workbook(buffer)
    assert wb.sheetnames == [SHEET_NAME]
