import math
import os

import openpyxl
import pytest
from openpyxl.cell.cell import MergedCell

from htmxl.compose import Workbook
from htmxl.compose.cell import Cell
from htmxl.compose.compat import BytesIO


class WriteTests:
    fixture_dir = ""
    template_file = ""
    expected_result_file = ""
    styles = []

    def load_result(self):
        return openpyxl.load_workbook(os.path.join(self.fixture_dir, self.expected_result_file))

    def load_source(self, parser):
        template_file = os.path.join(self.fixture_dir, self.template_file)
        wb = Workbook(parser=parser, styles=self.styles)
        wb.add_sheet_from_template_file(template_file)
        fileobj = BytesIO()
        wb.compose(fileobj)
        fileobj.seek(0)
        return openpyxl.load_workbook(fileobj)

    @pytest.mark.parametrize("parser", ["beautifulsoup", "lxml"])
    def test_equality(self, parser):
        result = self.load_source(parser)
        expected_result = self.load_result()

        result_worksheet = list(result.worksheets[0].rows)
        expected_worksheet = list(expected_result.worksheets[0].rows)

        self.display("Actual Result", result_worksheet)
        self.display("Expected Result", expected_worksheet)
        assert len(result_worksheet) == len(expected_worksheet)
        for result_row, expected_row in zip(result_worksheet, expected_worksheet):
            assert len(result_row) == len(expected_row)
            for result_cell, expected_cell in zip(result_row, expected_row):
                assert result_cell.value == expected_cell.value
                assert result_cell.style == expected_cell.style
                self.compare_width(result, result_cell, expected_result, expected_cell)

    def display(self, name, data):
        print(name)
        print("------")
        for row in data:
            row_values = []
            for cell in row:
                row_values.append(cell.value)

            print(" | ".join(str(v) for v in row_values if v is not None))
        print("------")
        print()

    def compare_width(self, result, result_cell, expected_result, expected_cell):
        if isinstance(result_cell, MergedCell):
            assert math.floor(
                result.worksheets[0].column_dimensions[Cell(result_cell.coordinate).col_ref].width
            ) == math.floor(
                expected_result.worksheets[0]
                .column_dimensions[Cell(expected_cell.coordinate).col_ref]
                .width
            )
        else:
            assert math.floor(
                result.worksheets[0].column_dimensions[result_cell.column_letter].width
            ) == math.floor(
                expected_result.worksheets[0].column_dimensions[expected_cell.column_letter].width
            )
